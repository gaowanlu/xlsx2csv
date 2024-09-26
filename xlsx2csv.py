import openpyxl
import csv
import os
import sys

# 将一个 .xlsx 文件的所有 sheet 转换为 CSV 文件
def xlsx_to_csv(xlsx_file, output_dir):
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
    
    # 获取文件名，不带扩展名
    file_name = os.path.splitext(os.path.basename(xlsx_file))[0]

    # 遍历工作簿中的所有 sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # 构建输出文件路径，文件名格式为 "原始文件名_表名.csv"
        csv_file = os.path.join(output_dir, f"{file_name}_{sheet_name}.csv")
        
        # 打开 CSV 文件进行写入
        with open(csv_file, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # 遍历 sheet 的所有行，将数据写入 CSV
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
        
        print(f"Sheet '{sheet_name}' from '{xlsx_file}' 已转换为 CSV: {csv_file}")

# 递归遍历文件夹，处理其中的所有 .xlsx 文件
def process_folder(input_dir, output_dir):
    # 确保输出目录存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 使用 os.walk() 递归遍历输入目录
    for root, dirs, files in os.walk(input_dir):
        for file_name in files:
            if file_name.endswith(".xlsx"):
                xlsx_file = os.path.join(root, file_name)
                
                # 构建相对路径结构以保持子文件夹层次
                relative_path = os.path.relpath(root, input_dir)
                output_subdir = os.path.join(output_dir, relative_path)
                
                # 确保子目录存在
                if not os.path.exists(output_subdir):
                    os.makedirs(output_subdir)
                
                print(f"正在处理文件: {xlsx_file}")
                xlsx_to_csv(xlsx_file, output_subdir)

# 主函数，通过命令行参数指定输入和输出路径
def main():
    if len(sys.argv) != 3:
        print("用法: python3 recursive_xlsx_to_csv.py <输入文件夹路径> <输出文件夹路径>")
        sys.exit(1)

    input_dir = sys.argv[1]
    output_dir = sys.argv[2]

    if not os.path.exists(input_dir):
        print(f"错误: 输入文件夹 '{input_dir}' 不存在。")
        sys.exit(1)

    process_folder(input_dir, output_dir)

if __name__ == "__main__":
    main()
